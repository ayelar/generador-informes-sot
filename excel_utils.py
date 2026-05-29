# ======================================
# SECCIÓN 1 - Carga del archivo Excel
# ======================================
                  
from openpyxl import load_workbook

def cargar_excel(excel_path):

    # Cargar workbook
    wb = load_workbook(excel_path, data_only=True)

    # Verificación defensiva
    if "Diccionario" not in wb.sheetnames:
        raise ValueError(
            "La hoja 'Diccionario' no existe en el archivo Excel"
        )

    ws_dic = wb["Diccionario"]

    return wb, ws_dic